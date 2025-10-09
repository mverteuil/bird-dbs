"""IOC World Bird Names database builder.

This utility streams IOC XML/XLSX data directly to SQLite database for
IOC (International Ornithological Committee) World Bird Names data.

The utility handles:
1. Streaming IOC XML taxonomy data to database
2. Processing multilingual XLSX translations to database
3. Creating optimized SQLite databases with indexes
4. Providing runtime lookups from database
"""

import contextlib
import xml.etree.ElementTree as ET
from collections.abc import Generator
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    import pandas as pd
    from openpyxl.worksheet.worksheet import Worksheet

from sqlalchemy import create_engine, delete, text
from sqlalchemy.orm import Session, sessionmaker
from sqlmodel import SQLModel

from ioc_reference_builder.ioc_models import (
    IOCLanguage,
    IOCMetadata,
    IOCSpecies,
    IOCTranslation,
)


class IOCDatabaseBuilder:
    """Builder for IOC World Bird Names SQLite databases."""

    def __init__(self, db_path: Path | str):
        """Initialize IOC database builder.

        Args:
            db_path: Required path to SQLite database
        """
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

        # Create database connection
        self.engine = create_engine(f"sqlite:///{self.db_path}")
        SQLModel.metadata.create_all(self.engine)
        self.session_local = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

        self._ioc_version = "unknown"

    @contextlib.contextmanager
    def get_db(self) -> Generator[Session, Any, None]:
        """Provide a database session for dependency injection."""
        db = self.session_local()
        try:
            yield db
        finally:
            db.close()

    def populate_from_files(
        self,
        xml_file: Path,
        avilistr_csv: Path,
        xlsx_file: Path | None = None,
    ) -> None:
        """Populate database from IOC XML and Avibase mapping files.

        Args:
            xml_file: Path to IOC XML file (required)
            avilistr_csv: Path to Avibase mapping CSV (required)
            xlsx_file: Path to IOC multilingual XLSX file (optional)
        """
        if not xml_file.exists():
            raise FileNotFoundError(f"XML file not found: {xml_file}")
        if not avilistr_csv.exists():
            raise FileNotFoundError(f"Avibase mapping file not found: {avilistr_csv}")

        print(f"Processing IOC XML: {xml_file}")
        print(f"Loading Avibase mapping: {avilistr_csv}")

        # Load Avibase mapping
        avibase_lookup = self._load_avibase_mapping(avilistr_csv)

        # Clear existing data
        with self.get_db() as session:
            session.execute(delete(IOCTranslation))
            session.execute(delete(IOCSpecies))
            session.execute(delete(IOCLanguage))
            session.execute(delete(IOCMetadata))
            session.commit()

        # Stream XML data directly to database
        species_count = self._stream_xml_to_database(xml_file, avibase_lookup)

        # Stream XLSX translations if provided
        translation_count = 0
        language_counts = {}
        if xlsx_file and xlsx_file.exists():
            print(f"Processing translations: {xlsx_file}")
            translation_count, language_counts = self._stream_xlsx_to_database(
                xlsx_file, avibase_lookup
            )

        # Store metadata
        self._store_metadata(species_count, translation_count, language_counts)

        # Create performance indexes
        self._create_performance_indexes()

        print(f"Database populated: {species_count} species, {translation_count} translations")

    def _load_avibase_mapping(self, avilistr_csv: Path) -> dict[str, str]:
        """Load Avibase ID mapping from CSV file.

        Args:
            avilistr_csv: Path to Avibase mapping CSV file

        Returns:
            Dictionary mapping scientific_name (lowercase) to avibase_id
        """
        import pandas as pd

        avilistr_data = pd.read_csv(avilistr_csv)
        avibase_lookup = {}

        for _, row in avilistr_data.iterrows():
            ioc_name = str(row["ioc_scientific_name"]).strip().lower()
            avibase_id = str(row["avibase_id"]).strip()
            avibase_lookup[ioc_name] = avibase_id

        print(f"  Loaded {len(avibase_lookup)} Avibase ID mappings")
        return avibase_lookup

    def _stream_xml_to_database(self, xml_file: Path, avibase_lookup: dict[str, str]) -> int:
        """Stream species data from IOC XML file directly to database.

        Args:
            xml_file: Path to IOC XML file
            avibase_lookup: Dictionary mapping scientific_name to avibase_id

        Returns:
            Number of species inserted
        """
        with self.get_db() as session:
            # Optimize SQLite for bulk insert
            session.execute(text("PRAGMA journal_mode = OFF"))
            session.execute(text("PRAGMA synchronous = OFF"))
            session.execute(text("PRAGMA temp_store = MEMORY"))
            session.execute(text("PRAGMA mmap_size = 268435456"))  # 256MB

            # Safe: IOC XML is from trusted source (IOC official database)
            tree = ET.parse(xml_file)  # nosemgrep
            root = tree.getroot()

            # Extract version
            self._ioc_version = root.get("version", "unknown")

            # Batch species data for bulk insert
            species_batch = []
            species_count = 0
            batch_size = 1000

            for order_elem in root.findall(".//order"):
                order_name = self._get_element_text(order_elem, "latin_name", "")

                for family_elem in order_elem.findall(".//family"):
                    family_latin = self._get_element_text(family_elem, "latin_name", "")

                    for genus_elem in family_elem.findall(".//genus"):
                        genus_name = self._get_element_text(genus_elem, "latin_name", "")
                        genus_authority = self._get_element_text(genus_elem, "authority", "")

                        for species_elem in genus_elem.findall("species"):
                            species_epithet = self._get_element_text(species_elem, "latin_name", "")
                            species_authority = self._get_element_text(
                                species_elem, "authority", ""
                            )
                            english_name = self._get_element_text(species_elem, "english_name", "")
                            breeding_regions = self._get_element_text(
                                species_elem, "breeding_regions", ""
                            )
                            breeding_subregions = self._get_element_text(
                                species_elem, "breeding_subregions", ""
                            )

                            # TODO: Detect and skip extinct species
                            # Check for daggers (†, ††) in english_name or extinct attribute

                            if genus_name and species_epithet and english_name:
                                scientific_name = f"{genus_name} {species_epithet}"
                                authority = species_authority or genus_authority

                                # Look up Avibase ID
                                avibase_id = avibase_lookup.get(scientific_name.lower())
                                if not avibase_id:
                                    # Skip species without Avibase ID mapping
                                    continue

                                species_batch.append(
                                    {
                                        "avibase_id": avibase_id,
                                        "scientific_name": scientific_name,
                                        "english_name": english_name,
                                        "order_name": order_name,
                                        "family": family_latin,
                                        "genus": genus_name,
                                        "species_epithet": species_epithet,
                                        "authority": authority,
                                        "breeding_regions": breeding_regions,
                                        "breeding_subregions": breeding_subregions,
                                    }
                                )
                                species_count += 1

                                # Insert batch when it reaches batch_size
                                if len(species_batch) >= batch_size:
                                    session.bulk_insert_mappings(IOCSpecies, species_batch)
                                    session.commit()
                                    species_batch = []
                                    print(f"  Inserted {species_count} species...")

            # Insert remaining species
            if species_batch:
                session.bulk_insert_mappings(IOCSpecies, species_batch)
                session.commit()

            # Re-enable normal SQLite settings
            session.execute(text("PRAGMA journal_mode = WAL"))
            session.execute(text("PRAGMA synchronous = NORMAL"))

        return species_count

    def _stream_xlsx_to_database(
        self, xlsx_file: Path, avibase_lookup: dict[str, str]
    ) -> tuple[int, dict[str, int]]:
        """Stream multilingual translations from XLSX directly to database.

        Args:
            xlsx_file: Path to IOC multilingual XLSX file
            avibase_lookup: Dictionary mapping scientific_name to avibase_id

        Returns:
            Tuple of (total translation count, language counts dict)
        """
        with self.get_db() as session:
            # Optimize for bulk insert
            self._set_bulk_insert_pragmas(session)

            wb = openpyxl.load_workbook(xlsx_file, read_only=True)
            ws = wb.active

            if ws is None:
                raise ValueError("No active worksheet in XLSX file")

            headers = self._extract_xlsx_headers(ws)
            lang_columns = self._map_language_columns(headers)
            ioc_col_idx = self._find_ioc_column_index(headers)

            # Stream translations
            translation_count, language_counts = self._process_xlsx_rows(
                ws, ioc_col_idx, lang_columns, avibase_lookup, session
            )

            # Insert language metadata
            self._insert_language_metadata(language_counts, session)

            # Re-enable normal settings
            self._restore_normal_pragmas(session)

        return translation_count, language_counts

    def _set_bulk_insert_pragmas(self, session: Session) -> None:
        """Set SQLite pragmas for optimized bulk insert."""
        session.execute(text("PRAGMA journal_mode = OFF"))
        session.execute(text("PRAGMA synchronous = OFF"))

    def _restore_normal_pragmas(self, session: Session) -> None:
        """Restore normal SQLite pragmas after bulk insert."""
        session.execute(text("PRAGMA journal_mode = WAL"))
        session.execute(text("PRAGMA synchronous = NORMAL"))

    def _extract_xlsx_headers(self, worksheet: "Worksheet") -> list[str]:
        """Extract headers from first row of worksheet."""
        headers = []
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) if cell else "" for cell in row]
            break
        return headers

    def _map_language_columns(self, headers: list[str]) -> dict[str, int]:
        """Map language headers to column indices."""
        lang_columns = {}
        excluded_headers = {"seq", "Order", "Family", "IOC_15.1", "English"}

        for i, header in enumerate(headers):
            if header and header not in excluded_headers:
                lang_code = self._map_language_to_code(header)
                if lang_code:
                    lang_columns[lang_code] = i

        return lang_columns

    def _find_ioc_column_index(self, headers: list[str]) -> int:
        """Find the IOC column index in headers."""
        if "IOC_15.1" not in headers:
            raise ValueError("IOC_15.1 column not found in XLSX")
        return headers.index("IOC_15.1")

    def _process_xlsx_rows(
        self,
        worksheet: "Worksheet",
        ioc_col_idx: int,
        lang_columns: dict[str, int],
        avibase_lookup: dict[str, str],
        session: Session,
    ) -> tuple[int, dict[str, int]]:
        """Process XLSX rows and insert translations in batches."""
        translation_batch = []
        translation_count = 0
        language_counts: dict[str, int] = {}
        batch_size = 1000

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row[ioc_col_idx]:
                continue

            scientific_name = str(row[ioc_col_idx]).strip()

            # Look up Avibase ID
            avibase_id = avibase_lookup.get(scientific_name.lower())
            if not avibase_id:
                # Skip species without Avibase ID mapping
                continue

            # Extract translations for each language
            for lang_code, col_idx in lang_columns.items():
                if col_idx < len(row) and row[col_idx]:
                    translated_name = str(row[col_idx]).strip()
                    if translated_name:
                        translation_batch.append(
                            {
                                "avibase_id": avibase_id,
                                "language_code": lang_code,
                                "common_name": translated_name,
                            }
                        )
                        translation_count += 1
                        language_counts[lang_code] = language_counts.get(lang_code, 0) + 1

                        # Insert batch when it reaches batch_size
                        if len(translation_batch) >= batch_size:
                            session.bulk_insert_mappings(IOCTranslation, translation_batch)
                            session.commit()
                            translation_batch = []
                            print(f"  Inserted {translation_count} translations...")

        # Insert remaining translations
        if translation_batch:
            session.bulk_insert_mappings(IOCTranslation, translation_batch)
            session.commit()

        return translation_count, language_counts

    def _insert_language_metadata(self, language_counts: dict[str, int], session: Session) -> None:
        """Insert language metadata into database."""
        language_names = self._get_language_names()
        for language_code, count in language_counts.items():
            language_name = language_names.get(language_code, language_code.upper())
            db_language = IOCLanguage(
                language_code=language_code,
                language_name=language_name,
                translation_count=count,
            )
            session.add(db_language)
        session.commit()

    def _store_metadata(
        self, species_count: int, translation_count: int, language_counts: dict[str, int]
    ) -> None:
        """Store metadata about the database."""
        from datetime import datetime

        with self.get_db() as session:
            metadata_entries = [
                IOCMetadata(key="ioc_version", value=self._ioc_version),
                IOCMetadata(key="created_at", value=datetime.utcnow().isoformat()),
                IOCMetadata(key="species_count", value=str(species_count)),
                IOCMetadata(key="translation_count", value=str(translation_count)),
                IOCMetadata(
                    key="languages_available", value=",".join(sorted(language_counts.keys()))
                ),
            ]

            for metadata in metadata_entries:
                session.add(metadata)
            session.commit()

    def _create_performance_indexes(self) -> None:
        """Create database indexes for optimal performance."""
        with self.get_db() as session:
            print("Creating performance indexes...")

            # Indexes for species table
            session.execute(
                text("""
                CREATE INDEX IF NOT EXISTS idx_species_family
                ON species(family)
            """)
            )

            session.execute(
                text("""
                CREATE INDEX IF NOT EXISTS idx_species_genus
                ON species(genus)
            """)
            )

            session.execute(
                text("""
                CREATE INDEX IF NOT EXISTS idx_species_order_family
                ON species(order_name, family)
            """)
            )

            session.execute(
                text("""
                CREATE INDEX IF NOT EXISTS idx_species_english_name
                ON species(english_name)
            """)
            )

            # Indexes for translations table
            # NOTE: Composite PK (avibase_id, language_code) automatically creates primary index
            # No need for separate idx_translations_avibase_language

            session.execute(
                text("""
                CREATE INDEX IF NOT EXISTS idx_translations_language_common
                ON translations(language_code, common_name)
            """)
            )

            session.commit()

    def _get_element_text(self, parent: ET.Element, tag_name: str, default: str) -> str:
        """Get text from child element with fallback."""
        elem = parent.find(tag_name)
        return elem.text.strip() if elem is not None and elem.text else default

    def _map_language_to_code(self, language_name: str) -> str | None:
        """Map language display name to ISO language code."""
        language_map = {
            "Catalan": "ca",
            "Chinese": "zh",
            "Chinese (Traditional)": "zh-TW",
            "Croatian": "hr",
            "Czech": "cs",
            "Danish": "da",
            "Dutch": "nl",
            "Finnish": "fi",
            "French": "fr",
            "German": "de",
            "Italian": "it",
            "Japanese": "ja",
            "Lithuanian": "lt",
            "Norwegian": "no",
            "Polish": "pl",
            "Portuguese (Lusophone)": "pt",
            "Portuguese (Portuguese)": "pt-PT",
            "Russian": "ru",
            "Serbian": "sr",
            "Slovak": "sk",
            "Spanish": "es",
            "Swedish": "sv",
            "Turkish": "tr",
            "Ukrainian": "uk",
            # Add more as needed
        }
        return language_map.get(language_name)

    def _get_language_names(self) -> dict[str, str]:
        """Get mapping of language codes to display names."""
        return {
            "ca": "Catalan",
            "zh": "Chinese",
            "zh-TW": "Chinese (Traditional)",
            "hr": "Croatian",
            "cs": "Czech",
            "da": "Danish",
            "nl": "Dutch",
            "fi": "Finnish",
            "fr": "French",
            "de": "German",
            "it": "Italian",
            "ja": "Japanese",
            "lt": "Lithuanian",
            "no": "Norwegian",
            "pl": "Polish",
            "pt": "Portuguese",
            "pt-PT": "Portuguese (Portuguese)",
            "ru": "Russian",
            "sr": "Serbian",
            "sk": "Slovak",
            "es": "Spanish",
            "sv": "Swedish",
            "tr": "Turkish",
            "uk": "Ukrainian",
        }

    def _parse_species_data(self, species_df: "pd.DataFrame") -> list[dict[str, str]]:
        """Parse species data from pandas DataFrame.

        Args:
            species_df: DataFrame with columns: Order, Family, Genus, Species, English name, Authority

        Returns:
            List of species dictionaries
        """

        species_list = []
        for _, row in species_df.iterrows():
            genus = str(row["Genus"]).strip()
            species_epithet = str(row["Species"]).strip()
            scientific_name = f"{genus} {species_epithet}"

            species_list.append(
                {
                    "scientific_name": scientific_name,
                    "english_name": str(row["English name"]).strip(),
                    "order_name": str(row["Order"]).strip(),
                    "family": str(row["Family"]).strip(),
                    "genus": genus,
                    "species_epithet": species_epithet,
                    "authority": (
                        str(row.get("Authority", "")).strip()
                        if row.get("Authority") is not None and str(row.get("Authority")).strip()
                        else None
                    ),
                }
            )
        return species_list

    def _parse_multilingual_data(self, translations_df: "pd.DataFrame") -> list[dict[str, str]]:
        """Parse multilingual translation data from pandas DataFrame.

        Args:
            translations_df: DataFrame with columns: Scientific name, LanguageCode, CommonName

        Returns:
            List of translation dictionaries
        """
        translations = []
        for _, row in translations_df.iterrows():
            common_name = str(row["CommonName"]).strip()
            language_code = str(row["LanguageCode"]).strip().lower()

            # Skip empty translations
            if not common_name:
                continue

            translations.append(
                {
                    "scientific_name": str(row["Scientific name"]).strip(),
                    "language_code": language_code,
                    "common_name": common_name,
                }
            )
        return translations

    def _map_avibase_ids(
        self, species_list: list[dict[str, str]], avilistr_data: "pd.DataFrame"
    ) -> list[dict[str, str]]:
        """Map Avibase IDs to species using Avilistr reference data.

        Args:
            species_list: List of species dictionaries
            avilistr_data: DataFrame with columns: scientific_name, avibase_id, ioc_scientific_name

        Returns:
            List of species dictionaries with avibase_id added
        """
        # Create lookup dict with case-insensitive keys
        avibase_lookup = {}
        for _, row in avilistr_data.iterrows():
            ioc_name = str(row["ioc_scientific_name"]).strip().lower()
            avibase_id = str(row["avibase_id"]).strip()
            avibase_lookup[ioc_name] = avibase_id

        # Map Avibase IDs to species
        mapped_species = []
        for species in species_list:
            scientific_name_lower = species["scientific_name"].lower()
            if scientific_name_lower in avibase_lookup:
                species_with_id = species.copy()
                species_with_id["avibase_id"] = avibase_lookup[scientific_name_lower]
                mapped_species.append(species_with_id)

        return mapped_species

    def _insert_species(self, species_list: list[dict[str, str]]) -> None:
        """Insert species into database.

        Args:
            species_list: List of species dictionaries with avibase_id
        """
        with self.get_db() as session:
            for species_data in species_list:
                # Check if species already exists
                existing = session.execute(
                    text("SELECT avibase_id FROM species WHERE avibase_id = :avibase_id"),
                    {"avibase_id": species_data["avibase_id"]},
                ).fetchone()

                if not existing:
                    species = IOCSpecies(**species_data)
                    session.add(species)

            session.commit()

    def _insert_translations(
        self, translations: list[dict[str, str]], avilistr_data: "pd.DataFrame"
    ) -> None:
        """Insert translations into database.

        Args:
            translations: List of translation dictionaries
            avilistr_data: DataFrame with Avibase ID mappings
        """
        # Create lookup dict with case-insensitive keys
        avibase_lookup = {}
        for _, row in avilistr_data.iterrows():
            ioc_name = str(row["ioc_scientific_name"]).strip().lower()
            avibase_id = str(row["avibase_id"]).strip()
            avibase_lookup[ioc_name] = avibase_id

        with self.get_db() as session:
            for trans_data in translations:
                scientific_name_lower = trans_data["scientific_name"].lower()
                if scientific_name_lower in avibase_lookup:
                    translation = IOCTranslation(
                        avibase_id=avibase_lookup[scientific_name_lower],
                        language_code=trans_data["language_code"],
                        common_name=trans_data["common_name"],
                    )
                    session.add(translation)

            session.commit()

    def get_statistics(self) -> dict[str, int]:
        """Get database statistics.

        Returns:
            Dictionary with species_count, translation_count, and language_count
        """
        with self.get_db() as session:
            species_count = session.execute(text("SELECT COUNT(*) FROM species")).scalar() or 0
            translation_count = (
                session.execute(text("SELECT COUNT(*) FROM translations")).scalar() or 0
            )
            language_count = (
                session.execute(
                    text("SELECT COUNT(DISTINCT language_code) FROM translations")
                ).scalar()
                or 0
            )

        return {
            "species_count": species_count,
            "translation_count": translation_count,
            "language_count": language_count,
        }

    # Query methods removed - use SpeciesDatabaseService for runtime queries
    # This builder is only responsible for creating the database from source files
